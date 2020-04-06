from tkinter import *
import tkinter.messagebox
import xlrd
import openpyxl
from tkinter import filedialog
import time
from openpyxl.reader.excel import load_workbook
import os
from fuzzywuzzy import fuzz
from fuzzywuzzy import process

def click():
    #try:

        file1="companyDomainDB.xlsx"
        readCDDB=xlrd.open_workbook(file1)
        file2=filedialog.askopenfilename(title="Select Excel file to generate email",filetypes=(("Excel file","*.xlsx"),("All files","*.*")))
        readIMP=xlrd.open_workbook(file2)
        sheetCDDBread=readCDDB.sheet_by_index(0)
        sheet2CDDBread=readCDDB.sheet_by_index(1)
        sheetIMPread=readIMP.sheet_by_index(0)
        

        book=openpyxl.Workbook()
        worksheet = book.active
        worksheet2=book.create_sheet("batch from LIn")
        
        worksheet.cell(row=1, column=1).value="Title"
        worksheet.cell(row=1, column=2).value="Name" 
        worksheet.cell(row=1, column=3).value="Surname"
        worksheet.cell(row=1, column=4).value="Email"
        worksheet.cell(row=1, column=5).value="Company"

        rows = 2
        rows2=1
        
        for r in range(sheetIMPread.nrows):
            companyNameMatch=[]
            for r3 in range(sheetCDDBread.nrows):
                if fuzz.token_set_ratio(sheetIMPread.cell_value(r,2),sheetCDDBread.cell_value(r3,0))>=90:
                    companyNameMatch.append(sheetCDDBread.cell_value(r3,0))
            if not companyNameMatch:
                    worksheet2.cell(row=rows2, column=1).value=sheetIMPread.cell_value(r,0)
                    worksheet2.cell(row=rows2, column=2).value=sheetIMPread.cell_value(r,1) 
                    worksheet2.cell(row=rows2, column=3).value=sheetIMPread.cell_value(r,2)
                    rows2=rows2+1                
            else:
                for r2 in range(sheetCDDBread.nrows):
                    if process.extractOne(sheetIMPread.cell_value(r,2), companyNameMatch)[0]==sheetCDDBread.cell_value(r2,0):
                        worksheet.cell(row=rows, column=1).value="Mr."
                        worksheet.cell(row=rows, column=2).value=sheetIMPread.cell_value(r,0)
                        worksheet.cell(row=rows, column=3).value=sheetIMPread.cell_value(r,1)
                        worksheet.cell(row=rows, column=5).value=sheetIMPread.cell_value(r,2)
                        for i in range(sheet2CDDBread.nrows):
                            if sheetIMPread.cell_value(r,0)==sheet2CDDBread.cell_value(i,0):
                                worksheet.cell(row=rows, column=1).value="Ms."
                                break
                            
                        if sheetCDDBread.cell_value(r2,2)=="John.Smith":
                        
                            worksheet.cell(row=rows, column=4).value= sheetIMPread.cell_value(r,0)+"."+sheetIMPread.cell_value(r,1)+"@"+sheetCDDBread.cell_value(r2,1)
                            
                        elif sheetCDDBread.cell_value(r2,2)=="John_Smith":
                            
                            worksheet.cell(row=rows, column=4).value=sheetIMPread.cell_value(r,0)+"_"+sheetIMPread.cell_value(r,1)+"@"+sheetCDDBread.cell_value(r2,1)
                            
                        elif sheetCDDBread.cell_value(r2,2)=="JohnSmith":
                            
                            worksheet.cell(row=rows, column=4).value= sheetIMPread.cell_value(r,0)+sheetIMPread.cell_value(r,1)+"@"+sheetCDDBread.cell_value(r2,1)
                            
                        elif sheetCDDBread.cell_value(r2,2)=="J.Smith":
                            
                            worksheet.cell(row=rows, column=4).value= sheetIMPread.cell_value(r,0)[0]+"."+sheetIMPread.cell_value(r,1)+"@"+sheetCDDBread.cell_value(r2,1)
                            
                            
                        elif sheetCDDBread.cell_value(r2,2)=="J_Smith":
                            
                            worksheet.cell(row=rows, column=4).value=sheetIMPread.cell_value(r,0)[0]+"_"+sheetIMPread.cell_value(r,1)+"@"+sheetCDDBread.cell_value(r2,1)
                            
                            
                        elif sheetCDDBread.cell_value(r2,2)=="JSmith":
                            
                            worksheet.cell(row=rows, column=4).value=sheetIMPread.cell_value(r,0)[0]+sheetIMPread.cell_value(r,1) +"@"+sheetCDDBread.cell_value(r2,1)
                           
                                                
                        elif sheetCDDBread.cell_value(r2,2)=="John.S":
                            
                            worksheet.cell(row=rows, column=4).value= sheetIMPread.cell_value(r,0)+"."+sheetIMPread.cell_value(r,1)[0]+"@"+sheetCDDBread.cell_value(r2,1)
                           
                            
                        elif sheetCDDBread.cell_value(r2,2)=="John_S":
                            
                            worksheet.cell(row=rows, column=4).value= sheetIMPread.cell_value(r,0)+"_"+sheetIMPread.cell_value(r,1)[0]+"@"+sheetCDDBread.cell_value(r2,1)
                            
                            
                        elif sheetCDDBread.cell_value(r2,2)=="JohnS":
                            
                            worksheet.cell(row=rows, column=4).value=sheetIMPread.cell_value(r,0)+sheetIMPread.cell_value(r,1)[0]+"@"+sheetCDDBread.cell_value(r2,1)
                           
                            
                        elif sheetCDDBread.cell_value(r2,2)=="John":
                            
                            worksheet.cell(row=rows, column=4).value= sheetIMPread.cell_value(r,0)+"@"+sheetCDDBread.cell_value(r2,1)
                            
                            
                        elif sheetCDDBread.cell_value(r2,2)=="Smith":
                            
                            worksheet.cell(row=rows, column=4).value= sheetIMPread.cell_value(r,1)+"@"+sheetCDDBread.cell_value(r2,1)
                        else:
                            worksheet.cell(row=rows, column=4).value= sheetIMPread.cell_value(r,0)+"."+sheetIMPread.cell_value(r,1)+"@"+sheetCDDBread.cell_value(r2,1)
                        rows += 1                 
                
            del companyNameMatch[:]
        book.save(os.path.basename(file2.split('.')[0])+time.strftime("%Y_%m_%d_%H_%M")+'.xlsx')
        tkinter.messagebox.showinfo("Email Generated","Task finished by generatting excel file called "+os.path.basename(file2.split('.')[0])+time.strftime("%Y_%m_%d_%H_%M")+'.xlsx')
    #except FileNotFoundError:
        #kidus=0
    #except Exception as e:
        #tkinter.messagebox.showerror("Something goes Wrong!",e)
				  
       
         
def click2():
    try:
        file1="companyDomainDB.xlsx"
        file3="oldDomainDB.xlsx"
        
        readIMP=xlrd.open_workbook(filedialog.askopenfilename(title="Select Domain Excel file to import domain to main DB",filetypes=(("Excel file","*.xlsx"),("All files","*.*"))))
        readDomainDB=xlrd.open_workbook(file3)
        
        sheetIMPread=readIMP.sheet_by_index(0)
        sheetDomainDB=readDomainDB.sheet_by_index(0)
        flag=0
        
        book =load_workbook(file1)
        sheet = book['CN+Domain+EF']
        rows=sheet.max_row+1
        maxrow_CDDB=sheet.max_row+1
        
        for r in range(0,sheetIMPread.nrows):
            for r2 in range(1,maxrow_CDDB):
                if sheetIMPread.cell_value(r,0)==sheet.cell(row=r2, column=1).value:
                    flag=flag+1
                    break
            if flag >0:
                tutututut=1
            elif flag==0:
                sheet.cell(row=rows, column=1).value=sheetIMPread.cell_value(r,0)
                sheet.cell(row=rows, column=2).value=sheetIMPread.cell_value(r,1)
                sheet.cell(row=rows, column=3).value="John.Smith"
                for r3 in range(sheetDomainDB.nrows):
                    if sheetIMPread.cell_value(r,1)==sheetDomainDB.cell_value(r3,0):
                        sheet.cell(row=rows, column=3).value=sheetDomainDB.cell_value(r3,1)
                        break
                        
                rows=rows+1            
                    
            flag=0
           
        book.save("companyDomainDB.xlsx")
    except FileNotFoundError:
        kidus=0
    except Exception as e:
        tkinter.messagebox.showerror("Something goes Wrong!",e)
def click3():
    try:
        file1="companyDomainDB.xlsx"
        
        readIMP=xlrd.open_workbook(filedialog.askopenfilename(title="Select Domain Excel file to update main DB",filetypes=(("Excel file","*.xlsx"),("All files","*.*"))))
        
        sheetIMPread=readIMP.sheet_by_index(0)

        book =load_workbook(file1)
        sheet = book['CN+Domain+EF']
        maxrow_CDDB=sheet.max_row+1
        
        for r in range(0,sheetIMPread.nrows):
            for r2 in range(1,maxrow_CDDB):
                if sheetIMPread.cell_value(r,0)==sheet.cell(row=r2, column=1).value:
                    sheet.cell(row=r2, column=2).value=sheetIMPread.cell_value(r,1)
                    sheet.cell(row=r2, column=3).value=sheetIMPread.cell_value(r,2)
        book.save("companyDomainDB.xlsx") 
    except FileNotFoundError:
        kidus=0
    except Exception as e:
        tkinter.messagebox.showerror("Something goes Wrong!",e)

def click4():
    try:
        fileBatch=filedialog.askopenfilename(title="Select Excel file to Exempt Sheet2 from Sheet1",filetypes=(("Excel file","*.xlsx"),("All files","*.*")))
        book =load_workbook(fileBatch)
        sheetBatch1 = book["Sheet1"]
        sheetBatch2=book["Sheet2"]
        sheetBatch3=book.create_sheet("Exempt S2 from S1")
        
        maxrow_Sheet1=sheetBatch1.max_row+1
        maxrow_Sheet2=sheetBatch2.max_row+1
        maxcolumn_Sheet1=sheetBatch1.max_column+1
        
        flag=0
        rows=1
        for r in range(1,maxrow_Sheet1):
            for r2 in range(1,maxrow_Sheet2):
                if sheetBatch1.cell(row=r, column=1).value==sheetBatch2.cell(row=r2, column=1).value:
                    flag+=1
            if flag==0:
                for r3 in range(1,maxcolumn_Sheet1):
                    sheetBatch3.cell(row=rows, column=r3).value=sheetBatch1.cell(row=r, column=r3).value
                rows+=1
            flag=0
            
        
        book.save(fileBatch)

    except FileNotFoundError:
        kidus=0
    except Exception as e:
        tkinter.messagebox.showerror("Something goes Wrong!",e)



def click5():
    try:
        fileBatch=filedialog.askopenfilename(title="Select Previously generated batch file",filetypes=(("Excel file","*.xlsx"),("All files","*.*")))
        book =load_workbook(fileBatch)
        sheetBatch = book["Sheet"]
        sheetBatch2=book.create_sheet("und from batch")
        readUndelivered=xlrd.open_workbook(filedialog.askopenfilename(title="Select Undelivered Excel file",filetypes=(("Excel file","*.xlsx"),("All files","*.*"))))
        sheetUndelivered=readUndelivered.sheet_by_index(0)
        
        maxrow_Batch=sheetBatch.max_row+1
        rows=2

        
        sheetBatch2.cell(row=1, column=1).value="Title"
        sheetBatch2.cell(row=1, column=2).value="Name" 
        sheetBatch2.cell(row=1, column=3).value="Surname"
        sheetBatch2.cell(row=1, column=4).value="Email"
        sheetBatch2.cell(row=1, column=5).value="Company"
        sheetBatch2.cell(row=1, column=6).value="Percent"
        sheetBatch2.cell(row=1, column=7).value="Domain in batch"
        sheetBatch2.cell(row=1, column=8).value="Domain in undelivered"
        
        batchDomains=[]
        undeliveredDomains=[]
        for r in range(0,sheetUndelivered.nrows):
            s=sheetUndelivered.cell_value(r,0)
            undeliveredDomains.append(s.split('@')[1])
            
        for r2 in range(2,maxrow_Batch):
            s2=sheetBatch.cell(row=r2, column=4).value
            batchDomains.append(s2.split('@')[1])

        for r in range(0,sheetUndelivered.nrows):
            for r2 in range(1,maxrow_Batch):
                if sheetUndelivered.cell_value(r,0)==sheetBatch.cell(row=r2, column=4).value:
                                        
                     sheetBatch2.cell(row=rows, column=1).value=sheetBatch.cell(row=r2, column=1).value
                     sheetBatch2.cell(row=rows, column=2).value=sheetBatch.cell(row=r2, column=2).value
                     sheetBatch2.cell(row=rows, column=3).value=sheetBatch.cell(row=r2, column=3).value
                     sheetBatch2.cell(row=rows, column=4).value=sheetBatch.cell(row=r2, column=4).value
                     sheetBatch2.cell(row=rows, column=5).value=sheetBatch.cell(row=r2, column=5).value
                     s3=sheetUndelivered.cell_value(r,0)
                     sheetBatch2.cell(row=rows, column=6).value=str((undeliveredDomains.count(s3.split('@')[1])/batchDomains.count(s3.split('@')[1]))*100)+"%"
                     sheetBatch2.cell(row=rows, column=7).value=batchDomains.count(s3.split('@')[1])
                     sheetBatch2.cell(row=rows, column=8).value=undeliveredDomains.count(s3.split('@')[1])
                     break
            rows=rows+1
        book.save(fileBatch)

    except FileNotFoundError:
        kidus=0
    except PermissionError as g:
        tkinter.messagebox.showerror("Something goes Wrong!","Please close the excel file which you are trying to upload.")
    except Exception as e:
        tkinter.messagebox.showerror("Something goes Wrong!",e)


def click6():
    os._exit(1)
def click7():
   tkinter.messagebox.showinfo("How to use the Software","Sharonat Email Generator Version 1.0\n Upcoming Update: Generate Separate Batches.\n\n******   Usage Instruction   ******\n\nGenerate Email:\n\n Input:\nExcel file(Scraped from LinkIn) with column 1: First name, Column 2: Surname and column 3: company name. \nOutput:\n New Excel file saved by time. sheet 1: Generated Emails, Sheet 2(batch from LIn): Rows could not find on Database.(This will be input for Import domain button.)\n\n Import Domain: \n\n Input:\nExcel file with column 1: Company Name and Column 2: Domain.\nOutput:\nNo visible out put(The Software will append the input to the main Excel.)\n\n UpdateDomain: \n\n Input:\nExcel file with column 1: Company Name, Column 2: Domain and Column 3: Email Format.\nOutput:\nNo visible out put(The Software will insert(Update) the input to the main Excel.) \n\nExcemption: \n\n Input:\nExcel file with has two sheets. \nSheet1: column 1:main list(it is possible to have multiple column but the operation use column one for excemption condition) and \nSheet2:has only has one Column  which has excemption list.\nOutput:\nThe sowftware will creat third Sheet(called Exempt S2 from S1) to show the result of excemption. \n\nBatch Vs Undelivered: \n\n Input:\nTwo Excel File: Excel 1:file Previously generated batch file.Which has\n column 1:Title,column 2:Name,column 3:Surname,column 4:Email and column 5:Company Name. \nExcel 2:underlivered emails Excel file.Has one Column with list of undelivered email. \nOutput:\nThe sowftware will creat Sheet(called Und from batch) to show the result of excemption. Which has und column 1:Title,column 2:Name,column 3:Surname,column 4:Email, column 5:Company Name, column 6:Percent of Underlivery, column 7:Number of Domains in Batch Excel and column 8:Number of Domains in Underlivered Excel. \n\nExit:\n\n To end the program.")
   
window=Tk()
window.title("Sharonat International");
window.configure(background="black")

photo=PhotoImage(file="sharonat.gif")
photo2=PhotoImage(file="info.png")
Label(window,image=photo,).grid(row=0, column=0, sticky=W)

Label(window,text="",bg="black", fg="black",font="none 12 bold").grid(row=1 , column=0,sticky=W)

Button(window,padx=16,pady=10,text="Generate Email", width=15, command=click).grid(row=2,column=0,sticky=W)
Button(window,padx=16,pady=10,text="Import Domain", width=15, command=click2).grid(row=2,column=0,sticky=E)

Label(window,text="",bg="black", fg="black",font="none 12 bold").grid(row=3 , column=0,sticky=W)

Button(window,padx=16,pady=10,text="Update Domain", width=15, command=click3).grid(row=4,column=0,sticky=W)
Button(window,padx=16,pady=10,text="Exemption", width=15, command=click4).grid(row=4,column=0,sticky=E)

Label(window,text="",bg="black", fg="black",font="none 12 bold" ).grid(row=5 , column=0,sticky=W)

Button(window,padx=16,pady=10,text="Batch Vs Undelivered", width=15, command=click5).grid(row=6,column=0,sticky=W)
Button(window,image=photo2, width=24, background="black",command=click7).grid(row=6,column=0)
Button(window,padx=16,pady=10,text="Exit", width=15, command=click6).grid(row=6,column=0,sticky=E)
window.mainloop()
